import io
from datetime import datetime, timedelta
from flask import Blueprint, jsonify, request, send_file
from db import get_db
from auth import api_admin_required

export_bp = Blueprint("export", __name__)


# ---------------------------------------------------------------------------
# GET /api/export/bills?period=today|daily|monthly|yearly|custom&from=&to=
# ---------------------------------------------------------------------------
@export_bp.route("/export/bills", methods=["GET"])
@api_admin_required
def export_bills():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        period = request.args.get("period", "monthly")
        db     = get_db()
        today  = datetime.now().date()

        if period == "today":
            today_str    = today.strftime("%Y-%m-%d")
            date_filter  = "b.bill_date = ?"
            params       = (today_str,)
            period_label = f"Today — {today.strftime('%d %b %Y')}"

        elif period == "daily":
            start = (today - timedelta(days=29)).strftime("%Y-%m-%d")
            end   = today.strftime("%Y-%m-%d")
            date_filter  = "b.bill_date BETWEEN ? AND ?"
            params       = (start, end)
            period_label = f"Last 30 Days ({start} to {end})"

        elif period == "yearly":
            start = f"{today.year - 4:04d}-01-01"
            end   = today.strftime("%Y-%m-%d")
            date_filter  = "b.bill_date BETWEEN ? AND ?"
            params       = (start, end)
            period_label = f"Last 5 Years ({start} to {end})"

        elif period == "custom":
            from_date = (request.args.get("from") or "").strip()
            to_date   = (request.args.get("to")   or "").strip()
            if not from_date or not to_date:
                return jsonify({"error": "from and to query params are required"}), 400
            try:
                datetime.strptime(from_date, "%Y-%m-%d")
                datetime.strptime(to_date,   "%Y-%m-%d")
            except ValueError:
                return jsonify({"error": "Dates must be in YYYY-MM-DD format"}), 400
            date_filter  = "b.bill_date BETWEEN ? AND ?"
            params       = (from_date, to_date)
            period_label = f"Custom Range ({from_date} to {to_date})"

        else:  # monthly — last 12 months
            total_m    = today.year * 12 + (today.month - 1) - 11
            start_year = total_m // 12
            start_mon  = (total_m % 12) + 1
            start      = f"{start_year:04d}-{start_mon:02d}-01"
            end        = today.strftime("%Y-%m-%d")
            date_filter  = "b.bill_date BETWEEN ? AND ?"
            params       = (start, end)
            period_label = f"Last 12 Months ({start} to {end})"

        rows = db.execute(f"""
            SELECT
                b.bill_date,
                strftime('%H:%M', b.created_at)                                    AS bill_time,
                b.bill_number,
                b.salesperson_name,
                b.payment_mode_type,
                COALESCE(bp.cash,  0)                                              AS cash,
                COALESCE(bp.card,  0)                                              AS card,
                COALESCE(bp.upi,   0)                                              AS upi,
                b.subtotal,
                b.total_discount,
                COALESCE(bi.gross_total, b.final_total) - b.final_total            AS round_off,
                b.final_total,
                b.advance_paid,
                b.remaining
            FROM bills b
            LEFT JOIN (
                SELECT bill_id,
                    SUM(CASE WHEN payment_method='Cash' THEN amount ELSE 0 END) AS cash,
                    SUM(CASE WHEN payment_method='Card' THEN amount ELSE 0 END) AS card,
                    SUM(CASE WHEN payment_method='UPI'  THEN amount ELSE 0 END) AS upi
                FROM bill_payments GROUP BY bill_id
            ) bp ON bp.bill_id = b.id
            LEFT JOIN (
                SELECT bill_id, SUM(final_amount) AS gross_total
                FROM bill_items GROUP BY bill_id
            ) bi ON bi.bill_id = b.id
            WHERE b.status != 'cancelled'
              AND {date_filter}
            ORDER BY b.bill_date ASC, b.created_at ASC
        """, params).fetchall()

        # ---- Build workbook ----
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"

        HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
        TOTAL_FILL  = PatternFill("solid", fgColor="D9E8FF")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
        TITLE_FONT  = Font(bold=True, size=12)
        TOTAL_FONT  = Font(bold=True, size=10)
        META_FONT   = Font(italic=True, color="666666", size=9)
        CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=False)
        LEFT        = Alignment(horizontal="left",   vertical="center")
        RIGHT       = Alignment(horizontal="right",  vertical="center")
        NUM_FMT     = '#,##0.00'

        COLS = [
            ("Date",             12, CENTER),
            ("Time",              7, CENTER),
            ("Bill No",          11, CENTER),
            ("Salesperson",      16, LEFT),
            ("Payment Type",     14, CENTER),
            ("Cash (₹)",         13, RIGHT),
            ("Card (₹)",         13, RIGHT),
            ("UPI (₹)",          13, RIGHT),
            ("Subtotal (₹)",     14, RIGHT),
            ("Discount (₹)",     13, RIGHT),
            ("Round Off (₹)",    13, RIGHT),
            ("Final Total (₹)",  15, RIGHT),
            ("Advance Paid (₹)", 15, RIGHT),
            ("Remaining (₹)",    14, RIGHT),
        ]
        NUM_COLS = len(COLS)

        def merge_title(row_num, value, font, height=20):
            last_col = get_column_letter(NUM_COLS)
            ws.merge_cells(f"A{row_num}:{last_col}{row_num}")
            c = ws.cell(row=row_num, column=1, value=value)
            c.font      = font
            c.alignment = CENTER
            ws.row_dimensions[row_num].height = height

        merge_title(1, "SHUBHAM NX — Sales Export", TITLE_FONT, height=22)
        merge_title(2, period_label, Font(size=10, color="444444"), height=16)
        merge_title(3, f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}", META_FONT, height=14)

        HDR_ROW = 5
        for col_i, (label, width, align) in enumerate(COLS, 1):
            c = ws.cell(row=HDR_ROW, column=col_i, value=label)
            c.font      = HEADER_FONT
            c.fill      = HEADER_FILL
            c.alignment = CENTER
            ws.column_dimensions[get_column_letter(col_i)].width = width
        ws.row_dimensions[HDR_ROW].height = 18

        DATA_START = HDR_ROW + 1
        col_totals = [0.0] * NUM_COLS

        for r_i, row in enumerate(rows, DATA_START):
            cash = float(row["cash"] or 0)
            card = float(row["card"] or 0)
            upi  = float(row["upi"]  or 0)

            values = [
                row["bill_date"],
                row["bill_time"] or "",
                row["bill_number"],
                row["salesperson_name"] or "",
                row["payment_mode_type"],
                cash,
                card,
                upi,
                float(row["subtotal"]       or 0),
                float(row["total_discount"] or 0),
                float(row["round_off"]      or 0),
                float(row["final_total"]    or 0),
                float(row["advance_paid"]   or 0),
                float(row["remaining"]      or 0),
            ]
            for col_i, (val, (_, _, align)) in enumerate(zip(values, COLS), 1):
                c = ws.cell(row=r_i, column=col_i, value=val)
                c.alignment = align
                if col_i >= 6:
                    c.number_format = NUM_FMT
                    col_totals[col_i - 1] += val

        TOTAL_ROW = DATA_START + len(rows)
        ws.merge_cells(f"A{TOTAL_ROW}:E{TOTAL_ROW}")
        label_cell = ws.cell(row=TOTAL_ROW, column=1, value="TOTAL")
        label_cell.font      = TOTAL_FONT
        label_cell.fill      = TOTAL_FILL
        label_cell.alignment = CENTER

        for col_i in range(6, NUM_COLS + 1):
            c = ws.cell(row=TOTAL_ROW, column=col_i, value=round(col_totals[col_i - 1], 2))
            c.font          = TOTAL_FONT
            c.fill          = TOTAL_FILL
            c.number_format = NUM_FMT
            c.alignment     = RIGHT
        ws.row_dimensions[TOTAL_ROW].height = 18

        ws.freeze_panes = f"A{DATA_START}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"sales_{period}_{today.strftime('%Y%m%d')}.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
